"""
삼창수산 & 장어만세 — 일일 매출 분석 자동 보고서 생성기
============================================================

매일 새벽 자동 실행되어:
  1) POS_DOWNLOADS 폴더의 어제 .xls 파일들을 모두 파싱
  2) 5개 지점 매출 분석 (수익률·인시매출·점심저녁비중·이상감지)
  3) HTML 메일 본문 생성
  4) 결과 출력 (Gmail 드래프트 자동 생성용 또는 SMTP 직접 발송)

실행:
  python daily_report_generator.py            # 어제 데이터로 보고서 생성
  python daily_report_generator.py 2026-05-04 # 특정 날짜
  python daily_report_generator.py --send     # 생성 + 즉시 SMTP 발송
"""
from __future__ import annotations
import argparse
import json
import re
import smtplib
import sys
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

try:
    import xlrd
except ImportError:
    print("[!] xlrd 가 설치되어 있지 않습니다.  pip install xlrd")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
DEFAULT_CONFIG = SCRIPT_DIR / "config.json"
HISTORY_FILE = SCRIPT_DIR / "sales_history.json"

STORE_COLOR = {
    '평택점': '#1F4E78', '오산점': '#2E75B6', '안성점': '#5B9BD5',
    '복대점': '#F59E0B', '율량점': '#16A34A', '삼창영어 서정지점': '#9333EA',
}


# ============================================================
# 유틸
# ============================================================
def yesterday_str():
    return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")


def won(n):
    if n is None or n == 0:
        return "-"
    return f"₩{int(n):,}"


def pct(n):
    if n is None:
        return "-"
    return f"{n*100:.1f}%"


def signed_pct(prev, curr):
    if not prev:
        return ("-", "flat")
    d = (curr - prev) / abs(prev)
    if abs(d) < 0.001:
        return ("0.0%", "flat")
    cls = "up" if d > 0 else "down"
    sign = "▲" if d > 0 else "▼"
    return (f"{sign} {abs(d)*100:.1f}%", cls)


# ============================================================
# .xls 파싱
# ============================================================
def parse_html_settlement(filepath: Path):
    """HTML 정산보고서 파싱 (복대점 등 POS별 형식)."""
    try:
        with open(str(filepath), 'rb') as f:
            raw = f.read()
        text = raw.decode('euc-kr', errors='replace')
        lines = [l.strip() for l in text.split('\n')]
        # 레이블-값 쌍 추출
        data = {}
        i = 0
        while i < len(lines):
            line = lines[i]
            if '<td align="center">' in line:
                label_m = re.search(r'<td[^>]*>([^<]+)</td>', line)
                if label_m:
                    label = label_m.group(1).strip()
                    # 다음 값 행들 찾기
                    vals = []
                    j = i + 1
                    while j < min(i + 4, len(lines)) and len(vals) < 2:
                        val_m = re.search(r'<td[^>]*>([\d,\-&nbsp;]+)</td>', lines[j])
                        if val_m:
                            v = val_m.group(1).replace(',','').replace('&nbsp;','').strip()
                            try: vals.append(int(v))
                            except: vals.append(0)
                        j += 1
                    if vals:
                        data[label] = vals
            i += 1
        # 핵심 값 추출
        total = data.get('총매출', [0, 0])[-1] if '총매출' in data else 0
        guests = data.get('객수', [0, 0])[0] if '객수' in data else 0
        avg = (total // guests) if guests else 0
        return {
            'date': None,
            'n_receipts': data.get('영수건수', [0])[0] if '영수건수' in data else 0,
            'lunch':  {'매출': 0, '객수': 0, '단가': 0},
            'dinner': {'매출': total, '객수': guests, '단가': avg},
            '_html_total': total,  # 점심/저녁 구분 없음 → 전체를 저녁으로 표시
        }
    except Exception as e:
        print(f"  HTML 정산 파싱 오류: {e}")
        return None


def parse_asp2_xlsx(filepath: Path):
    """
    ASP2 (복대점) 정산보고서 .xlsx 파싱.
    parse_okpos_xls와 동일한 dict 형식 반환.
    """
    try:
        import openpyxl
    except ImportError:
        print(f"  · {filepath.name}: openpyxl 미설치 — 스킵")
        return None
    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
    except Exception as e:
        print(f"  ASP2 .xlsx 파싱 실패: {filepath.name}: {e}")
        return None

    ws = wb[wb.sheetnames[0]]
    sheet_name = wb.sheetnames[0]
    m = re.search(r'(\d{4})_(\d{2})_(\d{2})', sheet_name)
    date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else None

    labels = {}
    for r in range(1, min(ws.max_row + 1, 60)):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str):
            labels[v.strip()] = r

    def cell(r, c):
        try:
            return ws.cell(row=r, column=c).value
        except:
            return None

    def to_int(v):
        if v is None: return 0
        try: return int(float(v))
        except: return 0

    n_receipts = to_int(cell(labels.get('영수건수', 5), 2))
    객수_total = to_int(cell(labels.get('객수', 6), 2))
    총매출 = to_int(cell(labels.get('총매출', 15), 3))

    lunch_sales = lunch_cust = 0
    dinner_sales = dinner_cust = 0
    hourly = {}

    time_header_r = None
    for r in range(1, min(ws.max_row + 1, 60)):
        v = cell(r, 5)
        if v and isinstance(v, str) and '시간대별' in v:
            time_header_r = r
            break

    if time_header_r:
        r = time_header_r + 1
        while r <= ws.max_row:
            time_str = cell(r, 5)
            if not time_str or not isinstance(time_str, str) or '~' not in time_str:
                next_v = cell(r + 1, 5)
                if not (isinstance(next_v, str) and '~' in next_v):
                    break
                r += 1
                continue
            try:
                hour = int(time_str.split(':')[0])
            except:
                r += 1
                continue
            cust = to_int(cell(r, 7))
            sales = to_int(cell(r, 8))
            is_lunch = (hour < 17)
            if is_lunch:
                lunch_sales += sales; lunch_cust += cust
            else:
                dinner_sales += sales; dinner_cust += cust
            if sales > 0:
                hourly[hour] = hourly.get(hour, 0) + sales
            r += 1

    if (lunch_sales + dinner_sales) == 0 and 총매출 > 0:
        dinner_sales = 총매출
        dinner_cust = 객수_total

    menu = {}
    menu_header_r = labels.get('상품명')
    if menu_header_r:
        r = menu_header_r + 1
        while r <= ws.max_row:
            name = cell(r, 1)
            qty = cell(r, 2)
            sales = cell(r, 3)
            if not name or not isinstance(name, str):
                if name is None and qty is None and sales is None:
                    break
                r += 1
                continue
            qty_i = to_int(qty)
            sales_i = to_int(sales)
            if sales_i <= 0:
                r += 1
                continue
            if '상차림' in name or '할인' in name or '적용' in name:
                r += 1
                continue
            total = lunch_sales + dinner_sales
            ratio_l = lunch_sales / total if total else 0
            l_part = int(sales_i * ratio_l)
            menu[name] = {
                'qty': qty_i,
                'sales': sales_i,
                'lunch': l_part,
                'dinner': sales_i - l_part,
            }
            r += 1

    avg_l = (lunch_sales // lunch_cust) if lunch_cust else 0
    avg_d = (dinner_sales // dinner_cust) if dinner_cust else 0

    return {
        'date': date,
        'n_receipts': n_receipts,
        'lunch':  {'매출': lunch_sales, '객수': lunch_cust, '단가': avg_l},
        'dinner': {'매출': dinner_sales, '객수': dinner_cust, '단가': avg_d},
        'menu': menu,
        'hourly': hourly,
        'turnover': {'tables': 0, 'visits': n_receipts, 'rate': 0},
    }


def parse_okpos_xls(filepath: Path):
    """OKPOS XLS 파일 파싱 — 형식 자동 감지 (영수증별매출상세현황 or 일기간(시간대별))."""
    # HTML 정산보고서 감지
    try:
        with open(str(filepath), 'rb') as f:
            magic = f.read(10)
        if magic.startswith(b'<!DOCTYPE') or magic.startswith(b'<html') or magic.startswith(b'\r\n<!'):
            print(f"  · {filepath.name}: HTML 정산보고서 형식 (일별 XLS 미수신)")
            return None
    except:
        pass
    try:
        wb = xlrd.open_workbook(str(filepath))
    except Exception as e:
        print(f"  XLS 파싱 실패: {filepath.name}: {e}")
        return None
    ws = wb.sheet_by_index(0)

    fmt = str(ws.cell_value(0, 0)) if ws.nrows > 0 else ''

    # 지원하지 않는 형식 (은행거래내역 등)
    if '영수증별매출상세현황' not in fmt and '일기간' not in fmt:
        return None  # 호출자에서 None 체크 필요

    # ─── 형식 1: 영수증별매출상세현황 ───────────────────────────────
    if '영수증별매출상세현황' in fmt:
        date = None
        for r in range(min(5, ws.nrows)):
            text = " ".join(str(ws.cell_value(r, c)) for c in range(ws.ncols))
            m = re.search(r'(\d{4}-\d{2}-\d{2})', text)
            if m:
                date = m.group(1); break

        sales_l = sales_d = 0
        cust_l = cust_d = 0
        n_receipts = set()
        menu = {}            # 상품명 → {qty, sales, lunch, dinner}
        hourly = {}          # 시간(int) → 매출액
        receipt_tables = {}  # 영수증번호 → 테이블명

        for r in range(6, ws.nrows - 1):  # 마지막 행 = 합계
            row = [ws.cell_value(r, c) for c in range(ws.ncols)]
            if not row[0]: continue
            receipt  = str(row[1])
            table    = str(row[3]).strip()   # 테이블명 (회전율용)
            time_str = str(row[5])   # 결제시각 HH:MM:SS
            product  = str(row[8])   # 상품명
            try: qty   = int(float(row[9]))  if row[9]  else 0
            except: qty = 0
            try: sales = int(float(row[15])) if row[15] else 0
            except: sales = 0

            n_receipts.add(receipt)
            if receipt not in receipt_tables:
                receipt_tables[receipt] = table
            try:
                hour = int(time_str.split(':')[0])
                is_lunch = (hour < 17)
            except:
                hour = -1
                is_lunch = False

            if is_lunch: sales_l += sales
            else:        sales_d += sales

            # 시간별 매출 집계
            if hour >= 0 and sales > 0:
                hourly[hour] = hourly.get(hour, 0) + sales

            # 메뉴 집계 (상차림·할인 제외)
            if product and '상차림' not in product and sales > 0:
                if product not in menu:
                    menu[product] = {'qty': 0, 'sales': 0, 'lunch': 0, 'dinner': 0}
                menu[product]['qty']   += qty
                menu[product]['sales'] += sales
                if is_lunch: menu[product]['lunch'] += sales
                else:        menu[product]['dinner'] += sales

            # 객수: 상차림N인 × 수량
            if '상차림' in product:
                m2 = re.search(r'(\d+)인', product)
                per_unit = int(m2.group(1)) if m2 else 1
                if is_lunch: cust_l += qty * per_unit
                else:        cust_d += qty * per_unit

        n_rec = len(n_receipts)
        # 회전율 계산 (포장 제외 숫자 테이블만)
        tbl_visits = {}
        for rec, tbl in receipt_tables.items():
            if tbl and '포장' not in tbl and tbl.replace(' ','').isdigit():
                tbl_visits[tbl] = tbl_visits.get(tbl, 0) + 1
        tbl_count = len(tbl_visits)
        tbl_total = sum(tbl_visits.values())
        turnover  = {'tables': tbl_count, 'visits': tbl_total,
                     'rate': round(tbl_total / tbl_count, 2) if tbl_count else 0}

        # 상차림 없는 매장(수산 직판장 등): 영수건수를 객수 대용으로 사용
        if cust_l == 0 and cust_d == 0 and n_rec > 0:
            # 점심/저녁 비율에 따라 영수건수 배분
            total_sales = sales_l + sales_d
            if total_sales > 0:
                ratio_l = sales_l / total_sales
                cust_l_fb = round(n_rec * ratio_l)
                cust_d_fb = n_rec - cust_l_fb
            else:
                cust_l_fb = cust_d_fb = 0
            avg_l = (sales_l // cust_l_fb) if cust_l_fb else 0
            avg_d = (sales_d // cust_d_fb) if cust_d_fb else 0
            return {
                'date': date,
                'n_receipts': n_rec,
                'lunch':  {'매출': sales_l, '객수': cust_l_fb, '단가': avg_l},
                'dinner': {'매출': sales_d, '객수': cust_d_fb, '단가': avg_d},
                'menu': menu,
                'hourly': hourly,
                'turnover': turnover,
                '_fallback_객수': True,
            }
        avg_l = (sales_l // cust_l) if cust_l else 0
        avg_d = (sales_d // cust_d) if cust_d else 0
        return {
            'date': date,
            'n_receipts': n_rec,
            'lunch':  {'매출': sales_l, '객수': cust_l, '단가': avg_l},
            'dinner': {'매출': sales_d, '객수': cust_d, '단가': avg_d},
            'menu': menu,
            'hourly': hourly,
            'turnover': turnover,
        }

    # ─── 형식 2: 일기간(시간대별) ─────────────────────────────────
    arr = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    date = None
    for row in arr[:6]:
        text = " ".join(str(v) for v in row)
        m = re.search(r'(\d{4}-\d{2}-\d{2})', text)
        if m:
            date = m.group(1); break

    lunch = {'매출': 0, '객수': 0, '단가': 0}
    dinner = {'매출': 0, '객수': 0, '단가': 0}
    for row in arr:
        label = str(row[1] if len(row) > 1 else '')
        try: cust = int(float(row[2])) if len(row) > 2 and row[2] else 0
        except: cust = 0
        try: sales = int(float(row[3])) if len(row) > 3 and row[3] else 0
        except: sales = 0
        try: avg = int(float(row[4])) if len(row) > 4 and row[4] else 0
        except: avg = 0
        if '점심' in label:
            lunch = {'매출': sales, '객수': cust, '단가': avg}
        elif '저녁' in label:
            dinner = {'매출': sales, '객수': cust, '단가': avg}

    return {'date': date, 'lunch': lunch, 'dinner': dinner}


def store_name_from_filename(filename: str, default: str = "?"):
    # 조합 파일 → 동일 지점(평택점)으로 매핑하여 합산
    if '송탄' in filename and '조합' not in filename:
        return '평택점'
    if '조합' in filename:
        return '평택점'
    stores = ['평택점', '오산점', '안성점', '복대점', '율량점', '삼창영어 서정지점', '서정지점']
    for s in stores:
        if s in filename or s.replace('점', '') in filename:
            return s if s != '서정지점' else '삼창영어 서정지점'
    return default


# ============================================================
# 분석
# ============================================================
def aggregate_stores(records: list[dict]) -> list[dict]:
    """records: [{store, date, lunch, dinner}, ...] → 지점별 집계."""
    by_store = {}
    for r in records:
        st = r['store']
        if st not in by_store:
            by_store[st] = {
                '지점': st, '점심매출': 0, '저녁매출': 0,
                '점심객수': 0, '저녁객수': 0, '일수': 0
            }
        a = by_store[st]
        a['점심매출'] += r['lunch']['매출']
        a['저녁매출'] += r['dinner']['매출']
        a['점심객수'] += r['lunch']['객수']
        a['저녁객수'] += r['dinner']['객수']
        a['일수'] += 1

    for a in by_store.values():
        a['총매출'] = a['점심매출'] + a['저녁매출']
        a['총객수'] = a['점심객수'] + a['저녁객수']
        a['점심비중'] = a['점심매출'] / a['총매출'] if a['총매출'] > 0 else 0
        a['객단가'] = a['총매출'] / a['총객수'] if a['총객수'] > 0 else 0

    return sorted(by_store.values(), key=lambda x: x['총매출'], reverse=True)


def load_history():
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_history(history):
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def update_history(history, date: str, records: list[dict]):
    """history[date][store] = {매출, 객수, 점심매출, 저녁매출}"""
    if date not in history:
        history[date] = {}
    for r in records:
        history[date][r['store']] = {
            '점심매출': r['lunch']['매출'],
            '저녁매출': r['dinner']['매출'],
            '점심객수': r['lunch']['객수'],
            '저녁객수': r['dinner']['객수'],
            '총매출': r['lunch']['매출'] + r['dinner']['매출'],
            '메뉴': r.get('menu', {}),
            '시간별': {str(k): v for k, v in r.get('hourly', {}).items()},
            '회전율': r.get('turnover', {}),
        }


# ============================================================
# HTML 보고서 생성
# ============================================================
def build_html_report(target_date: str, byStore: list[dict], history: dict, dashboard_url: str = ""):
    yesterday = (datetime.fromisoformat(target_date) - timedelta(days=1)).strftime("%Y-%m-%d")
    lastweek  = (datetime.fromisoformat(target_date) - timedelta(days=7)).strftime("%Y-%m-%d")

    total_sales = sum(s['총매출'] for s in byStore)
    total_lunch = sum(s['점심매출'] for s in byStore)
    total_dinner = sum(s['저녁매출'] for s in byStore)
    total_cust  = sum(s['총객수'] for s in byStore)
    avg_ticket  = total_sales / total_cust if total_cust else 0
    lunch_pct   = total_lunch / total_sales if total_sales else 0

    # 전일·전주 대비
    prev_total = sum(history.get(yesterday, {}).get(s['지점'], {}).get('총매출', 0) for s in byStore)
    week_total = sum(history.get(lastweek,  {}).get(s['지점'], {}).get('총매출', 0) for s in byStore)
    day_delta_text, day_delta_cls = signed_pct(prev_total, total_sales)
    wk_delta_text,  wk_delta_cls  = signed_pct(week_total, total_sales)

    # 이상 감지: 어제 대비 -15% 이상 떨어진 지점
    alerts = []
    for s in byStore:
        prev = history.get(yesterday, {}).get(s['지점'], {}).get('총매출', 0)
        if prev > 0 and (s['총매출'] - prev) / prev <= -0.15:
            drop = (prev - s['총매출']) / prev * 100
            alerts.append((s['지점'], s['총매출'], prev, drop))

    # KPI 색상 클래스
    def delta_html(text, cls):
        color = {'up': '#16a34a', 'down': '#dc2626', 'flat': '#64748b'}[cls]
        return f'<span style="color:{color};font-weight:600;font-size:13px">{text}</span>'

    rows_html = ""
    for s in byStore:
        prev = history.get(yesterday, {}).get(s['지점'], {}).get('총매출', 0)
        wkly = history.get(lastweek,  {}).get(s['지점'], {}).get('총매출', 0)
        d_txt, d_cls = signed_pct(prev, s['총매출'])
        w_txt, w_cls = signed_pct(wkly, s['총매출'])
        color = STORE_COLOR.get(s['지점'], '#64748b')
        rows_html += f"""
        <tr>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;font-weight:600;color:{color}">{s['지점']}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:right">{won(s['총매출'])}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:right">{won(s['점심매출'])}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:right">{won(s['저녁매출'])}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:center">{pct(s['점심비중'])}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:right">{won(s['객단가'])}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:center">{delta_html(d_txt, d_cls)}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:center">{delta_html(w_txt, w_cls)}</td>
        </tr>"""

    alert_html = ""
    if alerts:
        alert_items = "".join(
            f'<li style="margin-bottom:6px">'
            f'<b>{name}</b>: {won(curr)} (어제 {won(prev)}, <span style="color:#dc2626">▼ {drop:.1f}%</span>)</li>'
            for name, curr, prev, drop in alerts
        )
        alert_html = f"""
        <div style="background:#fee2e2;border:1px solid #fecaca;border-radius:8px;padding:14px 18px;margin:18px 0">
          <div style="font-weight:700;color:#991b1b;margin-bottom:8px">⚠ 매출 이상 감지 — 전일 대비 -15% 이상</div>
          <ul style="margin:0;padding-left:20px;color:#991b1b;font-size:13px">{alert_items}</ul>
        </div>"""

    # 메달 (1위 매출 지점)
    top_store = byStore[0] if byStore else None
    top_store_html = f'<b style="color:{STORE_COLOR.get(top_store["지점"], "#1F4E78")}">{top_store["지점"]} ({won(top_store["총매출"])})</b>' if top_store else "-"

    dashboard_link = f'<a href="{dashboard_url}" style="color:#fff;background:#1F4E78;padding:10px 22px;border-radius:6px;text-decoration:none;font-weight:600;display:inline-block">📊 전체 운영 대시보드 열기</a>' if dashboard_url else ""

    html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"></head>
<body style="margin:0;font-family:'Pretendard','Malgun Gothic','Apple SD Gothic Neo',Arial,sans-serif;background:#f1f5f9;color:#0f172a">
  <div style="max-width:680px;margin:0 auto;background:#fff">
    <div style="background:linear-gradient(135deg,#1F4E78,#2E75B6);color:#fff;padding:24px 28px">
      <div style="font-size:13px;opacity:.85;margin-bottom:4px">삼창수산 &amp; 장어만세 · 5개 지점</div>
      <div style="font-size:22px;font-weight:700">📊 {target_date} 일일 매출 분석 보고서</div>
    </div>

    <div style="padding:24px 28px">
      <table style="width:100%;border-collapse:collapse;margin-bottom:18px">
        <tr>
          <td style="padding:14px;background:#f8fafc;border-radius:8px;width:33%;text-align:center">
            <div style="color:#64748b;font-size:12px">총 매출</div>
            <div style="font-size:22px;font-weight:700;margin:4px 0">{won(total_sales)}</div>
            {delta_html(day_delta_text+' 전일대비', day_delta_cls)}
          </td>
          <td style="width:8px"></td>
          <td style="padding:14px;background:#f8fafc;border-radius:8px;width:33%;text-align:center">
            <div style="color:#64748b;font-size:12px">객단가 / 점심비중</div>
            <div style="font-size:22px;font-weight:700;margin:4px 0">{won(avg_ticket)}</div>
            <span style="font-size:13px;color:#64748b">점심 {lunch_pct*100:.1f}% · 객수 {total_cust:,}</span>
          </td>
          <td style="width:8px"></td>
          <td style="padding:14px;background:#f8fafc;border-radius:8px;width:33%;text-align:center">
            <div style="color:#64748b;font-size:12px">전주 동요일 대비</div>
            <div style="font-size:22px;font-weight:700;margin:4px 0">{won(total_sales)}</div>
            {delta_html(wk_delta_text+' 전주대비', wk_delta_cls)}
          </td>
        </tr>
      </table>

      {alert_html}

      <h3 style="margin:24px 0 8px;color:#1F4E78">🏪 지점별 비교</h3>
      <table style="width:100%;border-collapse:collapse;font-size:13px">
        <thead>
          <tr style="background:#f8fafc">
            <th style="padding:10px 12px;text-align:left;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">지점</th>
            <th style="padding:10px 12px;text-align:right;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">총매출</th>
            <th style="padding:10px 12px;text-align:right;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">점심</th>
            <th style="padding:10px 12px;text-align:right;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">저녁</th>
            <th style="padding:10px 12px;text-align:center;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">점심비중</th>
            <th style="padding:10px 12px;text-align:right;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">객단가</th>
            <th style="padding:10px 12px;text-align:center;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">전일</th>
            <th style="padding:10px 12px;text-align:center;color:#64748b;font-weight:600;border-bottom:1px solid #e2e8f0">전주</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>

      <div style="background:#f0f9ff;border-left:4px solid #2E75B6;padding:12px 16px;margin:22px 0;border-radius:4px">
        <div style="color:#1F4E78;font-weight:600;margin-bottom:6px">🥇 어제 매출 1위</div>
        <div>{top_store_html}</div>
      </div>

      <div style="text-align:center;margin:28px 0 12px">{dashboard_link}</div>

      <div style="border-top:1px solid #e2e8f0;padding-top:16px;margin-top:24px;color:#94a3b8;font-size:11px;text-align:center">
        본 보고서는 매일 새벽 자동 생성됩니다 · POS 다운로드 데이터 기준 · 6지점 종합
      </div>
    </div>
  </div>
</body></html>"""
    return html


# ============================================================
# Gmail 발송 (옵션)
# ============================================================
def send_gmail(html: str, target_date: str, cfg: dict):
    smtp_cfg = cfg.get("smtp")
    if not smtp_cfg:
        print("[i] config.json에 smtp 설정이 없어 발송을 건너뜁니다.")
        return False
    msg = MIMEMultipart('alternative')
    msg['Subject'] = f"[삼창수산] {target_date} 일일 매출 분석 보고서"
    msg['From']    = smtp_cfg['from']
    msg['To']      = ", ".join(smtp_cfg['to'])
    msg.attach(MIMEText("HTML 메일 클라이언트가 필요합니다.", 'plain'))
    msg.attach(MIMEText(html, 'html'))
    try:
        with smtplib.SMTP_SSL(smtp_cfg.get('host', 'smtp.gmail.com'),
                              smtp_cfg.get('port', 465)) as s:
            s.login(smtp_cfg['from'], smtp_cfg['app_password'])
            s.sendmail(smtp_cfg['from'], smtp_cfg['to'], msg.as_string())
        print(f"[✓] Gmail 발송 완료 → {', '.join(smtp_cfg['to'])}")
        return True
    except Exception as e:
        print(f"[✗] Gmail 발송 실패: {e}")
        return False


# ============================================================
# 메인
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="일일 매출 분석 보고서 자동 생성기")
    parser.add_argument("date", nargs="?", default=yesterday_str())
    parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    parser.add_argument("--send", action="store_true", help="HTML 생성 후 Gmail 즉시 발송")
    parser.add_argument("--out", help="HTML 보고서 저장 경로 (기본: SCRIPT_DIR/reports/)")
    args = parser.parse_args()

    cfg = {}
    if Path(args.config).exists():
        with open(args.config, 'r', encoding='utf-8') as f:
            cfg = json.load(f)

    download_dir = Path(cfg.get("download_folder", SCRIPT_DIR / "POS_DOWNLOADS"))
    if not download_dir.exists():
        print(f"[!] 다운로드 폴더 없음: {download_dir}")
        return 1

    print(f"\n┌{'─'*70}")
    print(f"│ 일일 매출 분석 보고서 — {args.date}")
    print(f"│ 데이터 폴더: {download_dir}")
    print(f"└{'─'*70}\n")

    # 1) 해당 날짜 .xls + .xlsx 모두 수집 (서브폴더와 루트 둘 다)
    # 파일명이 [평택점_2026-05-09.xls].xls 같은 변형도 매칭하기 위해 패턴 완화
    files = []
    for ext in ('xls', 'xlsx'):
        pat = f"*{args.date}*.{ext}"
        files.extend(download_dir.glob(pat))
        files.extend((download_dir / args.date).glob(pat))
    # 중복 제거 (같은 파일이 두 곳에서 검색될 가능성)
    seen_paths = set()
    files = [f for f in files if not (f.resolve() in seen_paths or seen_paths.add(f.resolve()))]
    if not files:
        print(f"[!] {args.date} 날짜의 .xls/.xlsx 파일이 없습니다.")
        return 2

    import hashlib
    seen_md5 = {}   # store → md5 (중복 감지용)
    store_data = {} # store → 합산 dict

    import os, hashlib as _hl
    _summary_path = os.environ.get('GITHUB_STEP_SUMMARY')
    def _write_summary(text):
        if _summary_path:
            with open(_summary_path, 'a', encoding='utf-8') as sf:
                sf.write(text + "\n")
        print(text)
    _write_summary(f"### 🐞 {args.date} 디버그 ({len(files)}개 파일)")
    _write_summary("| 파일명 | 크기 | 분류 store | MD5 |")
    _write_summary("|---|---|---|---|")
    for _f in files:
        _st = store_name_from_filename(_f.name)
        _md5 = _hl.md5(open(_f,'rb').read()).hexdigest()[:8]
        _write_summary(f"| `{_f.name}` | {_f.stat().st_size:,}B | `{_st}` | `{_md5}` |")
    for f in files:
        store = store_name_from_filename(f.name)
        print(f"  [DEBUG] processing {f.name} → store={store!r}, size={f.stat().st_size}")
        if store == '?':
            print(f"  · {f.name}: 지점 미확인 — 건너뜀")
            continue
        # MD5 중복 체크 (동일 파일이면 합산 건너뜀)
        fmd5 = hashlib.md5(open(f,'rb').read()).hexdigest()
        if seen_md5.get(store) == fmd5:
            print(f"  · {f.name}: {store}와 동일 파일(MD5 중복) — 건너뜀")
            continue
        seen_md5[store] = fmd5
        try:
            # 확장자별 파서 분기: .xlsx → ASP2 (복대점), .xls → OKPOS
            if f.suffix.lower() == '.xlsx':
                info = parse_asp2_xlsx(f)
            else:
                info = parse_okpos_xls(f)
            if info is None:
                _write_summary(f"- ⚠️ `{f.name}` → parse 결과 None (skip)")
                print(f"  · {f.name}: 지원하지 않는 형식 — 건너뜀")
                continue
            _write_summary(f"- 🔍 `{f.name}` parse: lunch ₩{info['lunch']['매출']:,}/{info['lunch']['객수']}객, dinner ₩{info['dinner']['매출']:,}/{info['dinner']['객수']}객")
            if store in store_data:
                # 동일 지점 두 번째 파일 → 합산
                prev = store_data[store]
                prev['lunch']['매출']  += info['lunch']['매출']
                prev['lunch']['객수']  += info['lunch']['객수']
                prev['dinner']['매출'] += info['dinner']['매출']
                prev['dinner']['객수'] += info['dinner']['객수']
                prev['n_receipts']     += info['n_receipts']
                # 메뉴 합산
                for prod, vals in info.get('menu', {}).items():
                    if prod not in prev.setdefault('menu', {}):
                        prev['menu'][prod] = {'qty':0,'sales':0,'lunch':0,'dinner':0}
                    for k in ('qty','sales','lunch','dinner'):
                        prev['menu'][prod][k] += vals.get(k,0)
                # 시간별 매출 합산
                for h, v in info.get('hourly', {}).items():
                    prev.setdefault('hourly', {})[h] = prev['hourly'].get(h, 0) + v
                # 회전율 합산
                pt = prev.setdefault('turnover', {'tables':0,'visits':0,'rate':0})
                nt = info.get('turnover', {'tables':0,'visits':0,'rate':0})
                pt['tables'] += nt.get('tables', 0)
                pt['visits'] += nt.get('visits', 0)
                pt['rate']    = round(pt['visits']/pt['tables'],2) if pt['tables'] else 0
                total = prev['lunch']['매출'] + prev['dinner']['매출']
                _write_summary(f"- ➕ `{f.name}` → {store} 누계 ₩{total:,}")
                print(f"  · {f.name}: {store}에 합산 → 누계 ₩{total:,}")
            else:
                store_data[store] = info
                total = info['lunch']['매출'] + info['dinner']['매출']
                _write_summary(f"- ✅ `{f.name}` → {store} 최초 ₩{total:,}")
                print(f"  · {store}: 매출 ₩{total:,}, 객수 {info['lunch']['객수']+info['dinner']['객수']}")
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"  ✗ {f.name} 파싱 실패: {e}")

    records = [{'store': st, 'date': args.date, **info} for st, info in store_data.items()]

    if not records:
        print("[!] 분석할 데이터가 없습니다.")
        return 2

    # 2) history 누적
    history = load_history()
    update_history(history, args.date, records)
    save_history(history)

    # 3) 분석 + HTML 생성
    byStore = aggregate_stores(records)
    dashboard_url = cfg.get("dashboard_url", "")
    html = build_html_report(args.date, byStore, history, dashboard_url)

    # 4) HTML 파일 저장
    out_dir = Path(args.out) if args.out else SCRIPT_DIR / "reports" / "일일"
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / f"보고서_{args.date}.html"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"\n[✓] HTML 보고서 저장: {report_path}")

    # 5) Gmail 발송
    if args.send:
        send_gmail(html, args.date, cfg)
    else:
        print(f"[i] --send 옵션을 추가하면 Gmail 자동 발송됩니다.")
        print(f"    또는 위 HTML 파일을 수동으로 메일 본문에 붙여 사용 가능합니다.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
